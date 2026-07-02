import pandas as pd
from openpyxl import load_workbook
import os

def actualizar_reporte_ventas(
    ruta_txt="C:/Users/Janus I5/Desktop/easygestion/descargas/ultimo_excel.txt",
    archivo_informe="C:/Users/Janus I5/Desktop/easygestion/Plantilla/Ventas_Marzo_Informe_Automatico.xlsx",
    hoja_datos="Datos"
):
    """
    Actualiza el informe de ventas diario:
    - Lee la ruta del archivo de ventas desde un archivo .txt
    - Agrupa ventas y facturas por asesor
    - Actualiza el archivo informe (Excel) con openpyxl:
        * Coincidencia exacta de nombre: actualiza ventas (col D) y facturas (col G)
        * Si no coincide pero los primeros 4 caracteres son dígitos (código),
          busca ese código en los datos de ventas, actualiza el nombre y los valores.
        * Si no se encuentra ni por nombre ni por código, coloca 0 en ventas y facturas.
    """
    # 1. Leer ruta del archivo de ventas desde el TXT
    try:
        with open(ruta_txt, "r", encoding="utf-8") as f:
            archivo_ventas = f.read().strip()
        print(f"Archivo de ventas a procesar: {archivo_ventas}")
    except FileNotFoundError:
        print(f"No se encontró el archivo {ruta_txt}. Verifica la ruta.")
        return
    except Exception as e:
        print(f"Error al leer {ruta_txt}: {e}")
        return

    if not os.path.exists(archivo_ventas):
        print(f"El archivo Excel de ventas no existe: {archivo_ventas}")
        return

    # 2. Cargar y procesar datos de ventas
    try:
        df = pd.read_excel(archivo_ventas)
    except Exception as e:
        print(f"Error al leer {archivo_ventas}: {e}")
        return

    columnas_necesarias = ["Asesor Nombre", "Subtotal Neto", "No EasySales"]
    for col in columnas_necesarias:
        if col not in df.columns:
            print(f"Error: La columna '{col}' no existe en el archivo de ventas.")
            return

    resumen = df.groupby("Asesor Nombre").agg(
        Ventas_Dia=("Subtotal Neto", "sum"),
        Numero_facturas=("No EasySales", "count")
    ).reset_index()

    resumen["Ventas_Dia"] = resumen["Ventas_Dia"].round().astype(int)
    resumen["Numero_facturas"] = resumen["Numero_facturas"].astype(int)

    datos_actualizados = {
        row["Asesor Nombre"]: (row["Ventas_Dia"], row["Numero_facturas"])
        for _, row in resumen.iterrows()
    }

    # 3. Función para extraer 4 dígitos iniciales
    def extraer_codigo_4d_inicial(nombre):
        nombre_str = str(nombre).strip()
        if len(nombre_str) >= 4 and nombre_str[:4].isdigit():
            return nombre_str[:4]
        return None

    # Mapeo código -> nombre correcto (desde datos_actualizados)
    codigo_a_nombre = {}
    for nombre in datos_actualizados.keys():
        cod = extraer_codigo_4d_inicial(nombre)
        if cod:
            codigo_a_nombre[cod] = nombre

    # 4. Actualizar informe con openpyxl
    if not os.path.exists(archivo_informe):
        print(f"El archivo informe {archivo_informe} no existe. Verifica.")
        return

    try:
        wb = load_workbook(archivo_informe)
        if hoja_datos not in wb.sheetnames:
            print(f"La hoja '{hoja_datos}' no existe en el informe.")
            return
        ws = wb[hoja_datos]
    except Exception as e:
        print(f"Error al abrir {archivo_informe}: {e}")
        return

    for row in range(2, ws.max_row + 1):
        asesor_celda = ws.cell(row, 2)  # Columna B = Asesor Comercial
        asesor_nombre_plantilla = asesor_celda.value
        if not asesor_nombre_plantilla:
            continue

        # Caso 1: Coincidencia exacta
        if asesor_nombre_plantilla in datos_actualizados:
            ventas, facturas = datos_actualizados[asesor_nombre_plantilla]
            ws.cell(row, 4, value=ventas)   # Col D
            ws.cell(row, 7, value=facturas) # Col G
            print(f"✓ Actualizado {asesor_nombre_plantilla}: Ventas={ventas}, Facturas={facturas}")
            continue

        # Caso 2: Buscar por código de 4 dígitos iniciales
        codigo = extraer_codigo_4d_inicial(asesor_nombre_plantilla)
        if codigo and codigo in codigo_a_nombre:
            nombre_correcto = codigo_a_nombre[codigo]
            ventas, facturas = datos_actualizados[nombre_correcto]

            # Actualizar nombre, ventas y facturas
            asesor_celda.value = nombre_correcto
            ws.cell(row, 4, value=ventas)
            ws.cell(row, 7, value=facturas)
            print(f" Nombre cambiado: '{asesor_nombre_plantilla}' → '{nombre_correcto}'")
            print(f"   Ventas={ventas}, Facturas={facturas}")
        else:
            # Caso 3: Sin ventas
            ws.cell(row, 4, value=0)
            ws.cell(row, 7, value=0)
            print(f"✗ Sin ventas: '{asesor_nombre_plantilla}' → 0")

    # Guardar cambios
    try:
        wb.save(archivo_informe)
        print(" Actualización completada. El informe ha sido guardado.")
    except Exception as e:
        print(f"Error al guardar {archivo_informe}: {e}")

# =========================
# EJECUTAR LA FUNCIÓN (si se corre el script directamente)
# =========================
if __name__ == "__main__":
    actualizar_reporte_ventas()